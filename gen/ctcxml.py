# encoding: utf-8
# -*- coding: uft-8 -*-

import time
from os import mkdir
from os.path import exists, join

import openpyxl

from app import get_config, write_file
from gen import render_template

template_name = 'ctcxml.html'
excel_path = unicode(get_config(__file__, 'excel_path'))
ROOT = get_config(__file__, 'root')
only_mark = bool(int(get_config(__file__, 'only_mark')))


def is_not_default_color(font):
    while_color = '00000000'
    return font.color.value != while_color


def get_svc_metadata(row):
    service_code = row[1]
    service_name = row[2]
    service_desc = row[3]
    if service_code.value:
        return {
            'code': service_code.value,
            'name': service_name.value,
            'desc': service_desc.value,
            'mark': True if is_not_default_color(service_code.font) else False
        }
    else:
        return None


def get_sheet_data(sheet):
    rows = sheet.rows
    col_max = sheet.max_column
    row_max = sheet.max_row

    # field header
    header = get_header(rows[0], 0, col_max)

    # data square
    col_min = 1
    row_min = 1

    # split request | response
    split_row_idx = get_split_row_idx(rows, row_min, row_max)

    # request
    req = get_square_data(header, rows, row_min, split_row_idx, col_min, col_max)

    # response
    resp = get_square_data(header, rows, split_row_idx, row_max, col_min, col_max)

    return req, resp


def get_header(row_header, col_min, col_max):
    ret = dict()
    field_name = None

    for col_idx in range(col_min, col_max):
        cell = row_header[col_idx]
        if cell.value:
            field_name = '_'.join(str(cell.value).split()).lower()
            ret[col_idx] = field_name
        elif field_name:
            ret[col_idx] = field_name
    return ret


def get_split_row_idx(rows, row_min, row_max):
    not_empty_row = 0
    for row_idx in range(row_min, row_max):
        row = rows[row_idx]
        if row[0].value:
            not_empty_row += 1
            if not_empty_row > 1:
                return row_idx

    raise Exception('excel format error, not req/resp format')


def get_square_data(header, rows, row_min, row_max, col_min, col_max):
    ret = {'name': rows[row_min][0].value, 'children': []}

    parent = ret
    prev_index0 = col_min

    for row_idx in range(row_min, row_max):
        row = rows[row_idx]

        row_obj = get_row_data(header, row, col_min, col_max)
        index0 = row_obj['index0']

        # same level
        if index0 == prev_index0:
            row_obj['parent'] = parent
            parent['children'].append(row_obj)

        # tab into
        elif index0 > prev_index0:
            parent = parent['children'][-1]

            row_obj['parent'] = parent
            parent['children'].append(row_obj)

        # tab out
        else:
            for times in range(0, prev_index0 - index0):
                parent = parent['parent']

            row_obj['parent'] = parent
            parent['children'].append(row_obj)

        prev_index0 = index0
    return ret


def get_row_data(header, row, col_min, col_max):
    ret = dict()
    for col_idx in range(col_min, col_max):
        cell = row[col_idx]
        if cell.value:
            if not ret:
                ret['index0'] = col_idx
            ret[header[col_idx]] = get_cell_value(cell.value)
    ret['children'] = []
    return ret


def get_cell_value(cell_value):
    if not isinstance(cell_value, basestring):
        return cell_value
    else:
        if '\n' in cell_value:
            return ' '.join([escape(value) for value in str(cell_value).splitlines()])
        else:
            return escape(cell_value)


def escape(cell_value):
    return cell_value \
        .replace('&', '&amp;') \
        .replace('<', '&lt;') \
        .replace('>', '&gt;') \
        .replace('"', '&quot;') \
        .replace('\'', '&apos;') \
        .strip()


def main():
    if not exists(ROOT):
        mkdir(ROOT)

    xml_root_dir = join(ROOT, 'xml')
    if not exists(xml_root_dir):
        mkdir(xml_root_dir)

    xml_dir = join(xml_root_dir, time.strftime('%Y%m%dH%H%M'))
    if not exists(xml_dir):
        mkdir(xml_dir)

    if not only_mark:
        xml_ref_dir = join(xml_dir, 'ref')
        mkdir(xml_ref_dir)

    wb = openpyxl.load_workbook(excel_path)

    for row in wb['Overview']:
        svc_metadata = get_svc_metadata(row)
        if svc_metadata:
            mark = svc_metadata['mark']
            if not mark and only_mark:
                continue
            try:
                sheet = wb.get_sheet_by_name(str(svc_metadata['code']))
                req, resp = get_sheet_data(sheet)
            except Exception as e:
                print e
            else:
                page = render_template(template_name, req=req, resp=resp, svc=svc_metadata)
                if page:
                    file_name = '{}.xml'.format(svc_metadata['code'])
                    write_file(file_name, page, xml_dir if mark else xml_ref_dir)

    print 'done'


if __name__ == '__main__':
    main()
