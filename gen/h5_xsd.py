# encoding: utf-8
# -*- coding: utf-8 -*-
from os.path import join

from openpyxl import load_workbook

from app import get_config, TMP_OUT
from gen.ctcxml import get_header
from gen.xsd import xls2xsd

EXCEL_PATH = unicode(get_config(__file__, 'excel_path'))
SHEET_NAME = get_config(__file__, 'sheet_name')


def get_cs_type(metadata):
    q = str(metadata).lower()
    if q == "dynamic" or q == 'code2':
        return "string"
    elif q == "int4" or q == "int10":
        return "int"
    elif q == "int20":
        return "long"
    elif q == "decimal2" or q == "decimal6":
        return 'decimal'
    elif q == "datetime" or q == "date":
        return 'DateTime'
    elif q == "boolean":
        return 'bool'
    else:
        return metadata


def gen_xls_xsd(excel_path, sheet_name):
    ''' 生成H5契约和XSD'''
    wb = load_workbook(excel_path)
    ws = wb.get_sheet_by_name(sheet_name)

    rows = ws.rows
    col_max = ws.max_column
    row_max = ws.max_row

    header = get_header(rows[0], 0, col_max)
    name_idx = None

    for row_idx in range(1, row_max):
        row = rows[row_idx]
        for col_idx in range(1, col_max):
            if header[col_idx] == 'name' and row[col_idx].value:
                name_idx = col_idx
            elif header[col_idx] == 'short_name':
                short_name = row[col_idx].value
                row[name_idx].value = short_name
            elif header[col_idx] == 'metadata':
                row[col_idx].value = get_cs_type(row[col_idx].value)

    wb_save_path = join(TMP_OUT, "{0}.xlsx".format(SHEET_NAME))
    wb.save(wb_save_path)
    xls2xsd(wb_save_path, SHEET_NAME)


if __name__ == '__main__':
    print EXCEL_PATH
    print SHEET_NAME

    gen_xls_xsd(EXCEL_PATH, SHEET_NAME)

    print 'done'
