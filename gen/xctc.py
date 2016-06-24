# encoding: utf-8
# -*- coding: uft-8 -*-

from gen.generate import Generator, CodeGenerateException
import openpyxl

__author__ = 'chengz'


class XContractGenerate(Generator):
    template_name = 'xcontract.html'

    def __init__(self, model):
        self.__model = model
        pass

    def _get_models(self):
        return self.__model

    pass


class ExcelFormatException(Exception):
    def __init__(self, *args, **kwargs):
        super(ExcelFormatException, self).__init__(args, kwargs)


def is_not_default_color(font):
    return font.color.value != '00000000'


def try_get_upd_svc(row):
    service_code = row[1]
    service_name = row[2]
    service_desc = row[3]
    if service_code.value and is_not_default_color(service_code.font):
        return {
            'code': service_code.value,
            'name': service_name.value,
            'desc': service_desc.value
        }
    else:
        return None


def get_sheet_data(sheet, svc_info):
    rows = sheet.rows
    col_max = sheet.max_column
    row_max = sheet.max_row

    # field header
    header = get_header(rows[0], 0, col_max)

    # data square
    col_min = 1
    row_min = 1

    # split request | response
    split_row_idx = get_split_row_idx(rows, row_min, row_max)

    # request
    req = get_square_data(header, rows, row_min, split_row_idx, col_min, col_max)

    # response
    resp = get_square_data(header, rows, split_row_idx, row_max, col_min, col_max)

    return {
        'base': {
            'name': svc_info['name'],
            'code': svc_info['code'],
            'desc': svc_info['desc']
        },
        'req': req,
        'resp': resp
    }


def get_header(row_header, col_min, col_max):
    ret = dict()
    field_name = None

    for col_idx in range(col_min, col_max):
        cell = row_header[col_idx]
        if cell.value:
            field_name = '_'.join(str(cell.value).split()).lower()
            ret[col_idx] = field_name
        elif field_name:
            ret[col_idx] = field_name
    return ret


def get_split_row_idx(rows, row_min, row_max):
    not_empty_row = 0
    for row_idx in range(row_min, row_max):
        row = rows[row_idx]
        if row[0].value:
            not_empty_row += 1
            if not_empty_row > 1:
                return row_idx

    raise ExcelFormatException()


def get_square_data(header, rows, row_min, row_max, col_min, col_max):
    ret = {'name': rows[row_min][0].value, 'children': []}

    parent = ret
    prev_index0 = col_min

    for row_idx in range(row_min, row_max):
        row = rows[row_idx]

        row_obj = get_row_data(header, row, col_min, col_max)
        index0 = row_obj['index0']

        # same level
        if index0 == prev_index0:
            row_obj['parent'] = parent
            parent['children'].append(row_obj)

        # tab into
        elif index0 > prev_index0:
            parent = parent['children'][-1]

            row_obj['parent'] = parent
            parent['children'].append(row_obj)

        # tab out
        else:
            for times in range(0, prev_index0 - index0):
                parent = parent['parent']

            row_obj['parent'] = parent
            parent['children'].append(row_obj)

        prev_index0 = index0
    return ret


def get_row_data(header, row, col_min, col_max):
    ret = dict()
    for col_idx in range(col_min, col_max):
        cell = row[col_idx]
        if cell.value:
            if not ret:
                ret['index0'] = col_idx
            ret[header[col_idx]] = get_cell_value(cell.value)
    ret['children'] = []
    return ret


def get_cell_value(cell_value):
    if not isinstance(cell_value, basestring):
        return cell_value
    if '\n' in cell_value:
        return '; '.join(str(cell_value).splitlines())

    return cell_value


def save_to_xml(upd_service_code, result):
    import time
    from os.path import exists
    from os import mkdir

    out_dir = 'd:/Users/chengz/Desktop/Contract/xml/{timestamp}'.format(timestamp=time.strftime('%Y%m%dH%H%M'))

    if not exists(out_dir):
        mkdir(out_dir)
    with open('{folder}/{code}.xml'.format(folder=out_dir, code=upd_service_code), 'w+') as fp:
        fp.write(result)


def main():
    from os.path import join, abspath
    file_path = join(abspath('..'), 'static/input/contract.xlsx')

    wb = openpyxl.load_workbook(file_path)

    for row in wb['Overview']:
        svc_info = try_get_upd_svc(row)
        if svc_info:
            try:
                sheet = wb.get_sheet_by_name(str(svc_info['code']))
                model = get_sheet_data(sheet, svc_info)
                result = XContractGenerate(model).generate()
            except ExcelFormatException:
                print 'error: ', 'excel format error'
            except CodeGenerateException as e:
                print 'error: ', e.message
            except Exception as e:
                print 'error: ', e.message
            else:
                if result:
                    save_to_xml(svc_info['code'], result)


if __name__ == '__main__':
    main()
    print 'done'
